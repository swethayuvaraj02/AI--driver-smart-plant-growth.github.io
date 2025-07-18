from flask import Flask, request, render_template_string, redirect, url_for, session
import pandas as pd
import openpyxl
import os
from werkzeug.utils import secure_filename
import secrets

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
app.config['SECRET_KEY'] = secrets.token_hex(16)

# Excel file for user credentials
USERS_EXCEL = 'users.xlsx'
if not os.path.exists(USERS_EXCEL):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"
    ws.append(["Username", "Password"])
    wb.save(USERS_EXCEL)

# Excel file for storing plant recommendations submitted by users
RECOMMENDATION_EXCEL = 'recommendations.xlsx'
if not os.path.exists(RECOMMENDATION_EXCEL):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Recommendations"
    # Columns: username + city + state + soil_type + recommended_plant + care_tips + light_needs + water_needs + image_url
    ws.append([
        "Username","City","State","Soil Type",
        "Recommended Plant","Care Tips","Light Needs","Water Needs","Image URL"
    ])
    wb.save(RECOMMENDATION_EXCEL)

# Load plant dataset
df = pd.read_excel("sample.xlsx")

# Login Template
LOGIN_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8" />
<meta name="viewport" content="width=device-width, initial-scale=1" />
<title>Login - Apartment Plant Recommender</title>
<style>
    body {
        margin: 0; padding: 0; 
        background: linear-gradient(135deg, #87A96B, #f1f8e9); 
        font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; color: #6b7280;
        display: flex; justify-content: center; align-items: center; height: 100vh;
    }
    .container {
        max-width: 400px; width: 100%; 
        background: #FFF6D2; 
        padding: 2rem; border-radius: 12px;
        box-shadow: 0 4px 12px rgb(0 0 0 / 0.1);
        text-align: center;
    }
    h2 {
        font-weight: 700; font-size: 2.5rem; color: #1b5e20; margin-bottom: 1rem;
    }
    form {
        display: flex; flex-direction: column; gap: 1.25rem;
    }
    label {
        font-weight: 600; text-align: left; color: #2e7d32;
    }
    input[type="text"], input[type="password"] {
        padding: 0.75rem 1rem; border: 1px solid #d1d5db; border-radius: 0.75rem; font-size: 1rem;
        transition: border-color 0.3s;background-color: #ECFFDC;
    }
    input[type="text"]:focus, input[type="password"]:focus {
        border-color: #355E3B; outline: none; 
    }
    input[type="submit"] {
        background-color: #2e7d32; color: white; font-weight: 700; font-size: 1rem;
        padding: 0.75rem; border: none; border-radius: 0.75rem; cursor: pointer; transition: background-color 0.3s;
    }
    input[type="submit"]:hover {
        background-color: #355E3B;
    }
    .message {
        margin-top: 1rem; color: #dc2626; font-weight: 600;
    }
    .link {
        margin-top: 1rem;
    }
    .link a {
        color: #355E3B;
        text-decoration: none;
    }
</style>
</head>
<body>
<div class="container" role="main" aria-label="Login form">
    <h2>Login</h2>
    <form method="POST" novalidate>
        <label for="username">Username</label>
        <input type="text" id="username" name="username" autocomplete="username" required />
        <label for="password">Password</label>
        <input type="password" id="password" name="password" autocomplete="current-password" required />
        <input type="submit" value="Log In" />
    </form>
    {% if message %}
        <div class="message" role="alert">{{ message }}</div>
    {% endif %}
    <div class="link">
        <a href="{{ url_for('register') }}">Don't have an account? Register here</a>
    </div>
</div>
</body>
</html>
"""

# Registration Template
REGISTER_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8" />
<meta name="viewport" content="width=device-width, initial-scale=1" />
<title>Register - Apartment Plant Recommender</title>
<style>
    body {
        margin: 0; padding: 0; 
        background: linear-gradient(135deg, #87A96B, #f1f8e9); 
        font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; color: #6b7280;
        display: flex; justify-content: center; align-items: center; height: 100vh;
    }
    .container {
        max-width: 400px; width: 100%; 
        background: #FFF6D2; 
        padding: 2rem; border-radius: 12px;
        box-shadow: 0 4px 12px rgb(0 0 0 / 0.1);
        text-align: center;
    }
    h2 {
        font-weight: 700; font-size: 2.5rem; color: #1b5e20; margin-bottom: 1rem;
    }
    form {
        display: flex; flex-direction: column; gap: 1.25rem;
    }
    label {
        font-weight: 600; text-align: left; color: #2e7d32;
    }
    input[type="text"], input[type="password"] {
        padding: 0.75rem 1rem; border: 1px solid #d1d5db; border-radius: 0.75rem; font-size: 1rem;
        transition: border-color 0.3s;background-color: #ECFFDC;
    }
    input[type="text"]:focus, input[type="password"]:focus {
        border-color: #2563eb; outline: none;
    }
    input[type="submit"] {
        background-color: #2e7d32; color: white; font-weight: 700; font-size: 1rem;
        padding: 0.75rem; border: none; border-radius: 0.75rem; cursor: pointer; transition: background-color 0.3s;
    }
    input[type="submit"]:hover {
        background-color: #355E3B;
    }
    .message {
        margin-top: 1rem; color: #dc2626; font-weight: 600;
    }
    .link {
        margin-top: 1rem;
    }
    .link a {
        color: #355E3B;
        text-decoration: none;
    }
</style>
</head>
<body>
<div class="container" role="main" aria-label="Registration form">
    <h2>Register</h2>
    <form method="POST" novalidate>
        <label for="username">Username</label>
        <input type="text" id="username" name="username" autocomplete="username" required />
        <label for="password">Password</label>
        <input type="password" id="password" name="password" autocomplete="new-password" required />
        <input type="submit" value="Register" />
    </form>
    {% if message %}
        <div class="message" role="alert">{{ message }}</div>
    {% endif %}
    <div class="link">
        <a href="{{ url_for('login') }}">Already have an account? Login here</a>
    </div>
</div>
</body>
</html>
"""

# Plant Recommendation Template
RECOMMENDATION_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8" />
<meta name="viewport" content="width=device-width, initial-scale=1" />
<title>Apartment Plant Recommender</title>
<style>
    body {
        margin: 0; padding: 2rem; background:  linear-gradient(135deg, #87A96B, #f1f8e9); 
        font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; color: #6b7280;
        display: flex; justify-content: center; min-height: 100vh;
    }
    .container {
        max-width: 600px; width: 100%; background: #FFF6D2; 
        padding: 2rem; border-radius: 0.75rem;
        box-shadow: 0 4px 12px rgb(0 0 0 / 0.1);
    }
    h2 {
        font-weight: 700; font-size: 2.5rem; color: #1b5e20; margin-bottom: 1rem; text-align: center;
    }
    form {
        display: flex; flex-direction: column; gap: 1rem;
        margin-bottom: 2rem;
    }
    label {
        font-weight: 600; color: #355E3B;
    }
    select, input[type="file"] {
        padding: 0.75rem 1rem; font-size: 1rem; border: 1px solid #d1d5db; border-radius: 0.75rem;
        transition: border-color 0.3s;background-color: #ECFFDC;
    }
    select:focus, input[type="file"]:focus {
        outline: none; border-color: #2563eb;
    }
    input[type="submit"] {
        background-color: #2e7d32; color: white; font-weight: 700; font-size: 1rem;
        padding: 0.75rem; border: none; border-radius: 0.75rem; cursor: pointer; transition: background-color 0.3s;
        margin-top: 1rem;
    }
    input[type="submit"]:hover {
        background-color: #355E3B;
    }
    .result {
        background: #ECFFDC; border-radius: 0.75rem; padding: 1.5rem; box-shadow: 0 1px 2px rgb(0 0 0 / 0.05);
        color: #111827;
    }
    .result h3 {
        font-weight: 700; margin-top: 0; margin-bottom: 0.5rem;
    }
    .result p {
        margin: 0.25rem 0;
    }
    img {
        max-width: 100%; border-radius: 0.5rem; margin-top: 1rem; border: 1px solid #d1d5db;
    }
    nav {
        display: flex; justify-content: flex-end; margin-bottom: 2rem;
    }
    nav a {
        text-decoration: none; color: #355E3B; font-weight: 600; padding: 0.5rem 1rem; border-radius: 0.5rem;
        border: 1px solid transparent; transition: background-color 0.3s, border-color 0.3s;
    }
    nav a:hover {
        background-color: #355E3B; color: white; border-color: #355E3B;
    }
</style>
</head>
<body>
<div class="container" role="main" aria-label="Plant Recommendation Form">

    <nav><a href="{{ url_for('logout') }}" aria-label="Logout">Logout</a></nav>

    <h2>Apartment-Friendly Plant Recommendation</h2>

    <form method="POST" enctype="multipart/form-data" novalidate>
        <label for="city">City</label>
        <select name="city" id="city" required aria-required="true">
            <option value="">--Select--</option>
            <option value="Chennai">Chennai</option>
            <option value="Mumbai">Mumbai</option>
            <option value="Bengaluru">Bengaluru</option>
            <option value="Delhi">Delhi</option>
            <option value="Kolkata">Kolkata</option>
            <option value="Jaipur">Jaipur</option>
            <option value="Hyderabad">Hyderabad</option>
            <option value="Ahmedabad">Ahmedabad</option>
            <option value="Lucknow">Lucknow</option>
            <option value="Bhopal">Bhopal</option>
            <option value="Patna">Patna</option>
            <option value="Surat">Surat</option>
            <option value="Nagpur">Nagpur</option>
            <option value="Visakhapatnam">Visakhapatnam</option>
            <option value="Itanagar">Itanagar</option>
            <option value="Guwahati">Guwahati</option>
            <option value="Raipur">Raipur</option>
            <option value="Panaji">Panaji</option>
            <option value="Gurugram">Gurugram</option>
            <option value="Shimla">Shimla</option>
            <option value="Ranchi">Ranchi</option>
            <option value="Thiruvananthapuram">Thiruvananthapuram</option>
            <option value="Imphal">Imphal</option>
            <option value="Shillong">Shillong</option>
            <option value="Aizawl">Aizawl</option>
            <option value="Kohima">Kohima</option>
            <option value="Bhubaneswar">Bhubaneswar</option>
            <option value="Gangtok">Gangtok</option>
            <option value="Agartala">Agartala</option>
            <option value="Dehradun">Dehradun</option>
        </select>

        <label for="state">State</label>
        <select name="state" id="state" required aria-required="true">
            <option value="">--Select--</option>
            <option value="Tamil Nadu">Tamil Nadu</option>
            <option value="Maharashtra">Maharashtra</option>
            <option value="Karnataka">Karnataka</option>
            <option value="Delhi">Delhi</option>
            <option value="West Bengal">West Bengal</option>
            <option value="Telangana">Telangana</option>
            <option value="Gujarat">Gujarat</option>
            <option value="Rajasthan">Rajasthan</option>
            <option value="Uttar Pradesh">Uttar Pradesh</option>
            <option value="Madhya Pradesh">Madhya Pradesh</option>
            <option value="Bihar">Bihar</option>
            <option value="Andhra Pradesh">Andhra Pradesh</option>
            <option value="Arunachal Pradesh">Arunachal Pradesh</option>
            <option value="Assam">Assam</option>
            <option value="Chhattisgarh">Chhattisgarh</option>
            <option value="Goa">Goa</option>
            <option value="Haryana">Haryana</option>
            <option value="Himachal Pradesh">Himachal Pradesh</option>
            <option value="Jharkhand">Jharkhand</option>
            <option value="Kerala">Kerala</option>
            <option value="Manipur">Manipur</option>
            <option value="Meghalaya">Meghalaya</option>
            <option value="Mizoram">Mizoram</option>
            <option value="Nagaland">Nagaland</option>
            <option value="Odisha">Odisha</option>
            <option value="Sikkim">Sikkim</option>
            <option value="Tripura">Tripura</option>
            <option value="Uttarakhand">Uttarakhand</option>
        </select>

        <label for="soil_image">Upload Soil Image</label>
        <input type="file" name="soil_image" id="soil_image" accept="image/*" required aria-required="true" />

        <input type="submit" value="Get Recommendation" />
    </form>

    {% if detected_soil %}
    <div class="result" role="region" aria-label="Soil Detection Result">
        <p><strong>Detected Soil Type:</strong> {{ detected_soil }}</p>
    </div>
    {% endif %}

    {% if recommendation %}
    <div class="result" role="region" aria-label="Plant Recommendation">
        <h3>🌿 Recommended Plant: {{ recommendation }}</h3>
        <p><strong>Care Tips:</strong> {{ care_tips }}</p>
        <p><strong>Light Needs:</strong> {{ light_needs }}</p>
        <p><strong>Water Needs:</strong> {{ water_needs }}</p>
        {% if image_url %}
            <img src="{{ image_url }}" alt="Image of {{ recommendation }}" />
        {% endif %}
    </div>
    {% endif %}
</div>
</body>
</html>
"""

# Dummy soil type classifier based on filename keywords
def dummy_soil_classifier(image_path):
    filename = os.path.basename(image_path).lower()
    if "clay" in filename:
        return "clay"
    elif "sandy" in filename:
        return "sandy"
    elif "loam" in filename:
        return "loamy"
    elif "black" in filename:
        return "black"
    elif "red" in filename:
        return "red"
    elif "alluvial" in filename:
        return "alluvial"
    else:
        return "loamy"  # default fallback

@app.route('/register', methods=['GET', 'POST'])
def register():
    message = ""
    if request.method == "POST":
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()

        if not username or not password:
            message = "Username and password are required."
            return render_template_string(REGISTER_HTML, message=message)

        wb = openpyxl.load_workbook(USERS_EXCEL)
        ws = wb.active
        # Check if username exists
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == username:
                message = "Username already exists. Please choose another."
                return render_template_string(REGISTER_HTML, message=message)

        ws.append([username, password])
        wb.save(USERS_EXCEL)
        message = "Registration successful! Please log in."
        return redirect(url_for('login'))

    return render_template_string(REGISTER_HTML, message=message)

@app.route('/login', methods=['GET', 'POST'])
def login():
    message = ""
    if request.method == "POST":
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()

        wb = openpyxl.load_workbook(USERS_EXCEL)
        ws = wb.active
        users = {}
        for row in range(2, ws.max_row + 1):
            u = ws.cell(row=row, column=1).value
            p = ws.cell(row=row, column=2).value
            users[u] = p

        if username in users and users[username] == password:
            session['logged_in'] = True
            session['username'] = username
            return redirect(url_for('index'))
        else:
            message = "Invalid username or password."
    return render_template_string(LOGIN_HTML, message=message)

@app.route('/', methods=['GET', 'POST'])
def index():
    if 'logged_in' not in session or not session['logged_in']:
        return redirect(url_for('login'))

    recommendation = care_tips = light_needs = water_needs = image_url = detected_soil = None
    username = session.get('username')

    if request.method == 'POST':
        city = request.form.get('city', '').strip().lower()
        state = request.form.get('state', '').strip().lower()

        if 'soil_image' not in request.files:
            return "Soil image is required.", 400

        file = request.files['soil_image']
        if file.filename == '':
            return "Please upload a valid soil image.", 400

        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        detected_soil = dummy_soil_classifier(filepath)
        soil_type = detected_soil

        matches = df[
            (df['city'].str.lower() == city) &
            (df['state'].str.lower() == state) &
            (df['soil_type'].str.lower() == soil_type)
        ]

        if not matches.empty:
            match = matches.iloc[0]
            recommendation = match['recommended_plant']
            care_tips = match['care_tips']
            light_needs = match['light_needs']
            water_needs = match['water_needs']
            image_url = match['image_url']
        else:
            recommendation = "No apartment-suitable plant found for these inputs."

        # Save recommendation with user data to Excel
        wb = openpyxl.load_workbook(RECOMMENDATION_EXCEL)
        ws = wb.active
        ws.append([
            username, city, state, soil_type,
            recommendation or "",
            care_tips or "",
            light_needs or "",
            water_needs or "",
            image_url or ""
        ])
        wb.save(RECOMMENDATION_EXCEL)

    return render_template_string(RECOMMENDATION_HTML,
                                  recommendation=recommendation,
                                  care_tips=care_tips,
                                  light_needs=light_needs,
                                  water_needs=water_needs,
                                  image_url=image_url,
                                  detected_soil=detected_soil)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

if __name__ == "_main_":
    app.run(debug=True)