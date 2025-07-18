from flask import Flask, request, render_template_string, redirect, url_for, session
import pandas as pd
import openpyxl
import os
from werkzeug.utils import secure_filename
import secrets
def get_season_from_month(month_name):
    month_name = month_name.strip().lower()
    if month_name in ['march', 'april', 'may']:
        return 'summer'
    elif month_name in ['june', 'july', 'august', 'september']:
        return 'monsoon'
    else:
        return 'winter'


app = Flask(_name_)
app.config['UPLOAD_FOLDER'] = 'uploads'
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
app.config['SECRET_KEY'] = secrets.token_hex(16)

# Excel file for user credentials
USERS_EXCEL = 'users.xlsx'
if not os.path.exists(USERS_EXCEL):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"
    ws.append(["Username", "Password"])
    wb.save(USERS_EXCEL)

# Excel file for storing plant recommendations submitted by users
RECOMMENDATION_EXCEL = 'recommendations.xlsx'
if not os.path.exists(RECOMMENDATION_EXCEL):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Recommendations"
    # Columns: username + city + state + soil_type + recommended_plant + care_tips + light_needs + water_needs + image_url
    ws.append([
        "Username","City","State","Soil Type",
        "Recommended Plant","Care Tips","Light Needs","Water Needs","Image URL"
    ])
    wb.save(RECOMMENDATION_EXCEL)

# Load plant dataset
df = pd.read_excel("sample.xlsx")

# Login Template
LOGIN_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8" />
<meta name="viewport" content="width=device-width, initial-scale=1" />
<title>Login - Apartment Plant Recommender</title>
<style>
    body {
        margin: 0; padding: 0; 
        background: linear-gradient(135deg, #E78B48 35%, #BE3D2A 100%); 
        font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; color: #6b7280;
        display: flex; justify-content: center; align-items: center; height: 100vh;
    }
    .container {
        max-width: 400px; width: 100%; 
        background: #FFDC7F; 
        padding: 2rem; border-radius: 12px;
        box-shadow: 0 4px 12px rgb(0 0 0 / 0.1);
        text-align: center;
    }
    h2 {
        font-weight: 700; font-size: 2.5rem; color: #521C0D; margin-bottom: 1rem;
    }
    form {
        display: flex; flex-direction: column; gap: 1.25rem;
    }
    label {
        font-weight: 600; text-align: left; color: #521C0D;
    }
    input[type="text"], input[type="password"] {
        padding: 0.75rem 1rem; border: 1px solid #d1d5db; border-radius: 0.75rem; font-size: 1rem;
        transition: border-color 0.3s;background-color: #ECFFDC;
    }
    input[type="text"]:focus, input[type="password"]:focus {
        border-color: #2563eb; outline: none;
    }
    input[type="submit"] {
        background-color: #521C0D; color: white; font-weight: 700; font-size: 1rem;
        padding: 0.75rem; border: none; border-radius: 0.75rem; cursor: pointer; transition: background-color 0.3s;
    }
    input[type="submit"]:hover {
        background-color: #521C0D;
    }
    .message {
        margin-top: 1rem; color: #521C0D; font-weight: 600;
    }
    .link {
        margin-top: 1rem;
    }
    .link a {
        color: #521C0D;
        text-decoration: none;
    }
    .leaf {
    position: fixed;
    top: -50px;
    width: 40px;
    height: 40px;
    background-image: url('https://cdn-icons-png.flaticon.com/512/766/766514.png');
    background-size: cover;
    animation: fall linear infinite;
    opacity: 0.8;
    z-index: 1;
    pointer-events: none;
}

@keyframes fall {
    0% {
        transform: translateY(-50px) translateX(0) rotate(0deg);
    }
    100% {
        transform: translateY(120vh) translateX(50px) rotate(360deg);
    }
}
</style>
</head>
<body>
<!-- Falling leaves (repeat as needed for more leaves) -->
<div class="leaf leaf1"></div>
<div class="leaf leaf2"></div>
<div class="leaf leaf3"></div>
<div class="leaf leaf4"></div>
<div class="leaf leaf5"></div>
<div class="container" role="main" aria-label="Login form">
    <h2>Login</h2>
    <form method="POST" novalidate>
        <label for="username">Username</label>
        <input type="text" id="username" name="username" autocomplete="username" required />
        <label for="password">Password</label>
        <input type="password" id="password" name="password" autocomplete="current-password" required />
        <input type="submit" value="Log In" />
    </form>
    {% if message %}
        <div class="message" role="alert">{{ message }}</div>
    {% endif %}
    <div class="link">
        <a href="{{ url_for('register') }}">Don't have an account? Register here</a>
    </div>
</div>


</body>
</html>
"""

# Registration Template
REGISTER_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8" />
<meta name="viewport" content="width=device-width, initial-scale=1" />
<title>Register - Apartment Plant Recommender</title>
<style>
    body {
        margin: 0; padding: 0; 
        background: linear-gradient(135deg, #E78B48 35%, #BE3D2A 100%); 
        font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; color: #6b7280;
        display: flex; justify-content: center; align-items: center; height: 100vh;
    }
    .container {
        max-width: 400px; width: 100%; 
        background: #FFDC7F; 
        padding: 2rem; border-radius: 12px;
        box-shadow: 0 4px 12px rgb(0 0 0 / 0.1);
        text-align: center;
    }
    h2 {
        font-weight: 700; font-size: 2.5rem; color: #521C0D; margin-bottom: 1rem;
    }
    form {
        display: flex; flex-direction: column; gap: 1.25rem;
    }
    label {
        font-weight: 600; text-align: left; color: #521C0D;
    }
    input[type="text"], input[type="password"] {
        padding: 0.75rem 1rem; border: 1px solid #d1d5db; border-radius: 0.75rem; font-size: 1rem;
        transition: border-color 0.3s;background-color: #ECFFDC;
    }
    input[type="text"]:focus, input[type="password"]:focus {
        border-color: #2563eb; outline: none;
    }
    input[type="submit"] {
        background-color: #521C0D; color: white; font-weight: 700; font-size: 1rem;
        padding: 0.75rem; border: none; border-radius: 0.75rem; cursor: pointer; transition: background-color 0.3s;
    }
    input[type="submit"]:hover {
        background-color: #521C0D;
    }
    .message {
        margin-top: 1rem; color: #521C0D; font-weight: 600;
    }
    .link {
        margin-top: 1rem;
    }
    .link a {
        color: #521C0D;
        text-decoration: none;
    }
    .leaf {
    position: fixed;
    top: -50px;
    width: 40px;
    height: 40px;
    background-image: url('https://cdn-icons-png.flaticon.com/512/766/766514.png');
    background-size: cover;
    animation: fall linear infinite;
    opacity: 0.8;
    z-index: 1;
    pointer-events: none;
}

@keyframes fall {
    0% {
        transform: translateY(-50px) translateX(0) rotate(0deg);
    }
    100% {
        transform: translateY(120vh) translateX(50px) rotate(360deg);
    }
}
</style>
</head>
<body>
<!-- Falling leaves (repeat as needed for more leaves) -->
<div class="leaf leaf1"></div>
<div class="leaf leaf2"></div>
<div class="leaf leaf3"></div>
<div class="leaf leaf4"></div>
<div class="leaf leaf5"></div>
<div class="container" role="main" aria-label="Registration form">
    <h2>Register</h2>
    <form method="POST" novalidate>
        <label for="username">Username</label>
        <input type="text" id="username" name="username" autocomplete="username" required />
        <label for="password">Password</label>
        <input type="password" id="password" name="password" autocomplete="new-password" required />
        <input type="submit" value="Register" />
    </form>
    {% if message %}
        <div class="message" role="alert">{{ message }}</div>
    {% endif %}
    <div class="link">
        <a href="{{ url_for('login') }}">Already have an account? Login here</a>
    </div>
</div>
<script>
// Generate multiple falling leaves with random positions and delays
for (let i = 0; i < 15; i++) {
    const leaf = document.createElement("div");
    leaf.className = "leaf";
    leaf.style.left = ${Math.random() * 100}vw; // Random left position
    leaf.style.animationDuration = ${5 + Math.random() * 5}s; // 5–10 seconds
    leaf.style.animationDelay = ${Math.random() * 5}s; // Random start delay
    leaf.style.zIndex = 1;
    document.body.appendChild(leaf);
}
</script>
</body>
</html>
"""

# Plant Recommendation Template
RECOMMENDATION_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8" />
<meta name="viewport" content="width=device-width, initial-scale=1" />
<title>Apartment Plant Recommender</title>
<style>
    body {
        margin: 0;
        padding: 2rem;
        background: linear-gradient(135deg, #E78B48 35%, #BE3D2A 100%), url('https://www.transparenttextures.com/patterns/leaf.png');
        background-blend-mode: overlay;
        font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
        color: #521C0D;
        display: flex;
        justify-content: center;
        min-height: 100vh;
        overflow: auto;
    }

    .container {
        max-width: 650px;
        width: 100%;
        background: #FFDC7F;
        padding: 2rem 2.5rem;
        border-radius: 1rem;
        box-shadow: 0 8px 20px rgba(34, 70, 50, 0.15);
        border: 2px dashed #A2C29A;
        backdrop-filter: blur(3px);
        position: relative;
        z-index: 2;
    }

    h2 {
        font-weight: 800;
        font-size: 2.5rem;
        color: #521C0D;
        margin-bottom: 1rem;
        text-align: center;
        letter-spacing: 0.05em;
    }

    form {
        display: flex;
        flex-direction: column;
        gap: 1rem;
        margin-bottom: 2rem;
    }

    label {
        font-weight: 600;
        color: #521C0D;
    }

    select, input[type="file"] {
        padding: 0.75rem 1rem;
        font-size: 1rem;
        border: 2px solid #521C0D;
        border-radius: 0.75rem;
        background-color: #F3FFE2;
        transition: all 0.3s;
    }

    select:focus, input[type="file"]:focus {
        outline: none;
        border-color: #521C0D;
        box-shadow: 0 0 0 2px rgba(102, 187, 106, 0.3);
    }

    input[type="submit"] {
        background-color: #521C0D;
        color: white;
        font-weight: 700;
        font-size: 1.05rem;
        padding: 0.85rem;
        border: none;
        border-radius: 0.75rem;
        cursor: pointer;
        transition: background-color 0.3s, transform 0.2s;
        display: flex;
        align-items: center;
        justify-content: center;
        gap: 0.5rem;
    }

    input[type="submit"]::before {
        content: "🌿";
        font-size: 1.2rem;
    }

    input[type="submit"]:hover {
        background-color: #521C0D;
        transform: scale(1.02);
    }

    .result {
        background: #F0FFF0;
        border-radius: 0.75rem;
        padding: 1.5rem;
        box-shadow: 0 4px 8px rgba(0, 50, 0, 0.05);
        border-left: 5px solid #521C0D;
        color: #1B4332;
        padding
    }

    .result h3 {
        font-weight: 700;
        margin-top: 0;
        margin-bottom: 0.5rem;
        font-size: 1.4rem;
    }

    .result p {
        margin: 0.25rem 0;
        font-size: 1rem;
    }

    img {
        max-width: 100%;
        border-radius: 0.5rem;
        margin-top: 1rem;
        border: 1px solid #d1d5db;
        box-shadow: 0 2px 4px rgba(0,0,0,0.05);
    }

    nav {
        display: flex;
        justify-content: flex-end;
        margin-bottom: 2rem;
    }

    nav a {
        text-decoration: none;
        color: #33691E;
        font-weight: 600;
        padding: 0.6rem 1.2rem;
        border-radius: 0.5rem;
        border: 1px solid #A5D6A7;
        background-color: #E8F5E9;
        transition: all 0.3s;
    }

    nav a:hover {
        background-color: #43A047;
        color: white;
        border-color: #43A047;
    }

    /* Animated falling leaves */
    .leaf {
    position: fixed;
    top: -50px;
    width: 40px;
    height: 40px;
    background-image: url('https://pngimg.com/uploads/autumn_leaves/autumn_leaves_PNG3613.png');
    background-size: cover;
    animation: fall linear infinite;
    opacity: 0.8;
    z-index: 1;
    pointer-events: none;
    
}

@keyframes fall {
    0% {
        transform: translateY(-50px) translateX(0) rotate(0deg);
    }
    100% {
        transform: translateY(120vh) translateX(50px) rotate(360deg);
    }
}

</style>
</head>
<body>
<!-- Falling leaves (repeat as needed for more leaves) -->
<div class="leaf leaf1"></div>
<div class="leaf leaf2"></div>
<div class="leaf leaf3"></div>
<div class="leaf leaf4"></div>
<div class="leaf leaf5"></div>


<div class="container" role="main" aria-label="Plant Recommendation Form">

    <nav><a href="{{ url_for('logout') }}" aria-label="Logout">Logout</a></nav>

    <h2>Plant Recommendation</h2>

    <form method="POST" enctype="multipart/form-data" novalidate>
        <label for="city">City</label>
        <select name="city" id="city" required aria-required="true">
            <option value="">--Select--</option>
            <option value="Chennai">Chennai</option>
            <option value="Mumbai">Mumbai</option>
            <option value="Bengaluru">Bengaluru</option>
            <option value="Delhi">Delhi</option>
            <option value="Kolkata">Kolkata</option>
            <option value="Jaipur">Jaipur</option>
            <option value="Hyderabad">Hyderabad</option>
            <option value="Ahmedabad">Ahmedabad</option>
            <option value="Lucknow">Lucknow</option>
            <option value="Bhopal">Bhopal</option>
            <option value="Patna">Patna</option>
            <option value="Surat">Surat</option>
            <option value="Nagpur">Nagpur</option>
            <option value="Visakhapatnam">Visakhapatnam</option>
            <option value="Itanagar">Itanagar</option>
            <option value="Guwahati">Guwahati</option>
            <option value="Raipur">Raipur</option>
            <option value="Panaji">Panaji</option>
            <option value="Gurugram">Gurugram</option>
            <option value="Shimla">Shimla</option>
            <option value="Ranchi">Ranchi</option>
            <option value="Thiruvananthapuram">Thiruvananthapuram</option>
            <option value="Imphal">Imphal</option>
            <option value="Shillong">Shillong</option>
            <option value="Aizawl">Aizawl</option>
            <option value="Kohima">Kohima</option>
            <option value="Bhubaneswar">Bhubaneswar</option>
            <option value="Gangtok">Gangtok</option>
            <option value="Agartala">Agartala</option>
            <option value="Dehradun">Dehradun</option>
        </select>

        <label for="state">State</label>
        <select name="state" id="state" required aria-required="true">
            <option value="">--Select--</option>
            <option value="Tamil Nadu">Tamil Nadu</option>
            <option value="Maharashtra">Maharashtra</option>
            <option value="Karnataka">Karnataka</option>
            <option value="Delhi">Delhi</option>
            <option value="West Bengal">West Bengal</option>
            <option value="Telangana">Telangana</option>
            <option value="Gujarat">Gujarat</option>
            <option value="Rajasthan">Rajasthan</option>
            <option value="Uttar Pradesh">Uttar Pradesh</option>
            <option value="Madhya Pradesh">Madhya Pradesh</option>
            <option value="Bihar">Bihar</option>
            <option value="Andhra Pradesh">Andhra Pradesh</option>
            <option value="Arunachal Pradesh">Arunachal Pradesh</option>
            <option value="Assam">Assam</option>
            <option value="Chhattisgarh">Chhattisgarh</option>
            <option value="Goa">Goa</option>
            <option value="Haryana">Haryana</option>
            <option value="Himachal Pradesh">Himachal Pradesh</option>
            <option value="Jharkhand">Jharkhand</option>
            <option value="Kerala">Kerala</option>
            <option value="Manipur">Manipur</option>
            <option value="Meghalaya">Meghalaya</option>
            <option value="Mizoram">Mizoram</option>
            <option value="Nagaland">Nagaland</option>
            <option value="Odisha">Odisha</option>
            <option value="Sikkim">Sikkim</option>
            <option value="Tripura">Tripura</option>
            <option value="Uttarakhand">Uttarakhand</option>
        </select>

        <label for="Month">Month</label>
        <select name="Month" id="Month" required aria-required="true">

            <option value="">--Select--</option>
            <option value="January">January</option>
            <option value="February">February</option>
            <option value="March">March</option>
            <option value="April">April</option>
            <option value="May">May</option>
            <option value="June">June</option>
            <option value="July">July</option>
            <option value="August">August</option>
            <option value="September">September</option>
            <option value="October">October</option>
            <option value="November">November</option>
            <option value="December">December</option>

        </select>

        <label for="soil_image">Upload Soil Image</label>
        <input type="file" name="soil_image" id="soil_image" accept="image/*" required aria-required="true" />

        <input type="submit" value="Get Recommendation" />
    </form>

    {% if detected_soil %}
    <div class="result" role="region" aria-label="Soil Detection Result" style="padding=2px">
        <p><strong>Detected Soil Type:</strong> {{ detected_soil }}</p>
    </div>
    {% endif %}

    {% if season %}
    <div class="result" role="region" aria-label="Soil Detection Result">
        <p><strong>Detected Season:</strong> {{ season }}</p>
    </div>
    {% endif %}




    {% if recommendation %}
    <div class="result" role="region" aria-label="Plant Recommendation">
        <h3>🌿 Recommended Plant: {{ recommendation }}</h3>
        <p><strong>Care Tips:</strong> {{ care_tips }}</p>
        <p><strong>Light Needs:</strong> {{ light_needs }}</p>
        <p><strong>Water Needs:</strong> {{ water_needs }}</p>
        {% if image_url %}
            <img src="{{ image_url }}" alt="Image of {{ recommendation }}" />
        {% endif %}
    </div>
    {% endif %}
</div>
<script>
// Generate multiple falling leaves with random positions and delays
for (let i = 0; i < 15; i++) {
    const leaf = document.createElement("div");
    leaf.className = "leaf";
    leaf.style.left = ${Math.random() * 100}vw; // Random left position
    leaf.style.animationDuration = ${5 + Math.random() * 5}s; // 5–10 seconds
    leaf.style.animationDelay = ${Math.random() * 5}s; // Random start delay
    leaf.style.zIndex = 1;
    document.body.appendChild(leaf);
}
</script>

</body>
</html>
"""
#detecting season
def get_season_from_month_name(month_name):
    month_name = month_name.strip().lower()
    if month_name in ['march', 'april', 'may']:
        return 'summer'
    elif month_name in ['june', 'july', 'august', 'september']:
        return 'monsoon'
    elif month_name in ['october', 'november', 'december', 'january', 'february']:
        return 'winter'
    return 'unknown'

# Dummy soil type classifier based on filename keywords
def dummy_soil_classifier(image_path):
    filename = os.path.basename(image_path).lower()
    if "clay" in filename:
        return "clay"
    elif "sandy" in filename:
        return "sandy"
    elif "loam" in filename:
        return "loamy"
    elif "black" in filename:
        return "black"
    elif "red" in filename:
        return "red"
    elif "alluvial" in filename:
        return "alluvial"
    else:
        return "loamy"  # default fallback

@app.route('/register', methods=['GET', 'POST'])
def register():
    message = ""
    if request.method == "POST":
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
       


        if not username or not password:
            message = "Username and password are required."
            return render_template_string(REGISTER_HTML, message=message)

        wb = openpyxl.load_workbook(USERS_EXCEL)
        ws = wb.active
        # Check if username exists
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == username:
                message = "Username already exists. Please choose another."
                return render_template_string(REGISTER_HTML, message=message)

        ws.append([username, password])
        wb.save(USERS_EXCEL)
        message = "Registration successful! Please log in."
        return redirect(url_for('login'))

    return render_template_string(REGISTER_HTML, message=message)

@app.route('/login', methods=['GET', 'POST'])
def login():
    message = ""
    if request.method == "POST":
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()

        wb = openpyxl.load_workbook(USERS_EXCEL)
        ws = wb.active
        users = {}
        for row in range(2, ws.max_row + 1):
            u = ws.cell(row=row, column=1).value
            p = ws.cell(row=row, column=2).value
            users[u] = p

        if username in users and users[username] == password:
            session['logged_in'] = True
            session['username'] = username
            return redirect(url_for('index'))
        else:
            message = "Invalid username or password."
    return render_template_string(LOGIN_HTML, message=message)

@app.route('/', methods=['GET', 'POST'])
@app.route('/', methods=['GET', 'POST'])
def index():
    if 'logged_in' not in session or not session['logged_in']:
        return redirect(url_for('login'))

    recommendation = care_tips = light_needs = water_needs = image_url = detected_soil = season = None
    username = session.get('username')

    if request.method == 'POST':
        city = request.form.get('city', '').strip().lower()
        state = request.form.get('state', '').strip().lower()
        month_selected = request.form.get('Month', '').strip()

        # Detect season based on selected month
        season = get_season_from_month(month_selected)

        if 'soil_image' not in request.files:
            return "Soil image is required.", 400

        file = request.files['soil_image']
        if file.filename == '':
            return "Please upload a valid soil image.", 400

        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        detected_soil = dummy_soil_classifier(filepath)
        soil_type = detected_soil

        # Filter DataFrame with all conditions
        if 'season' in df.columns:
            df['season'] = df['season'].fillna('').astype(str)
            matches = df[
                (df['city'].str.lower() == city) &
                (df['state'].str.lower() == state) &
                (df['soil_type'].str.lower() == soil_type) &
                (df['season'].str.lower() == season.lower())
            ]
        else:
            matches = df[
                (df['city'].str.lower() == city) &
                (df['state'].str.lower() == state) &
                (df['soil_type'].str.lower() == soil_type)
            ]

        if not matches.empty:
            match = matches.iloc[0]
            recommendation = match['recommended_plant']
            care_tips = match['care_tips']
            light_needs = match['light_needs']
            water_needs = match['water_needs']
            image_url = match['image_url']
        else:
            recommendation = "No apartment-suitable plant found for these inputs."

        # Save the result
        wb = openpyxl.load_workbook(RECOMMENDATION_EXCEL)
        ws = wb.active
        ws.append([
            username, city, state, soil_type,
            recommendation or "", care_tips or "", light_needs or "",
            water_needs or "", image_url or ""
        ])
        wb.save(RECOMMENDATION_EXCEL)

    return render_template_string(RECOMMENDATION_HTML,
                                  recommendation=recommendation,
                                  care_tips=care_tips,
                                  light_needs=light_needs,
                                  water_needs=water_needs,
                                  image_url=image_url,
                                  detected_soil=detected_soil,
                                  season=season)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

if _name_ == "_main_":
    app.run(debug=True)
